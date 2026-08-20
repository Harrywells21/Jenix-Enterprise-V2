from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from db import get_db, Machine, AuditLog, CveScan, CveFinding, SessionLocal
from auth import get_current_user, require_operator, User
from cve_scanner import run_cve_scan
from datetime import datetime
import json

router = APIRouter(prefix="/cve", tags=["cve"])

# Cache scan results in memory
_scan_cache = {}

@router.post("/scan/{machine_id}")
async def trigger_cve_scan(
    machine_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    m = db.query(Machine).filter(Machine.id == machine_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Machine not found")

    # For local machine — run directly
    # For remote — would send via agent (future enhancement)
    log = AuditLog(
        machine_id = machine_id,
        user_id    = current_user.id,
        action     = "cve_scan",
        detail     = f"CVE scan triggered by {current_user.name}",
        status     = "ok"
    )
    db.add(log); db.commit()

    # Run scan in background
    triggered_name = current_user.name
    triggered_id   = current_user.id

    def _do_scan():
        result = run_cve_scan(max_packages=30)
        _scan_cache[machine_id] = result

        # Persist to DB so results survive restarts and build real history
        session = SessionLocal()
        try:
            scan_row = CveScan(
                machine_id          = machine_id,
                triggered_by_id     = triggered_id,
                triggered_by_name   = triggered_name,
                scanned_at          = datetime.utcnow(),
                packages_scanned    = result.get("packages_scanned", 0),
                vulnerable_packages = result.get("vulnerable_packages", 0),
                total_vulns         = result.get("total_vulns", 0),
                critical            = result.get("critical", 0),
                high                = result.get("high", 0),
                risk_level          = result.get("risk_level", "LOW"),
            )
            session.add(scan_row)
            session.commit()
            session.refresh(scan_row)

            for pkg_result in result.get("results", []):
                for v in pkg_result.get("vulns", []):
                    finding = CveFinding(
                        scan_id  = scan_row.id,
                        package  = pkg_result.get("package", ""),
                        version  = pkg_result.get("version", ""),
                        cve_id   = v.get("id", "Unknown"),
                        summary  = v.get("summary", ""),
                        severity = v.get("severity", "UNKNOWN"),
                        url      = v.get("url", ""),
                    )
                    session.add(finding)
            session.commit()
        except Exception as e:
            print(f"[cve] failed to persist scan results: {e}")
            session.rollback()
        finally:
            session.close()

    background_tasks.add_task(_do_scan)
    return {"ok": True, "message": "CVE scan started — check back in 30 seconds"}

@router.get("/results/{machine_id}")
def get_cve_results(
    machine_id: int,
    db: Session = Depends(get_db),
    _:  User    = Depends(get_current_user)
):
    if machine_id not in _scan_cache:
        return {"scanned": False,
                "message": "No scan results yet. Run a scan first."}
    return {"scanned": True, **_scan_cache[machine_id]}

@router.get("/summary")
def cve_summary(db: Session = Depends(get_db),
                _:  User    = Depends(get_current_user)):
    total_critical = sum(
        r.get("critical", 0) for r in _scan_cache.values()
    )
    total_high = sum(
        r.get("high", 0) for r in _scan_cache.values()
    )
    machines_scanned = len(_scan_cache)
    return {
        "machines_scanned": machines_scanned,
        "total_critical":   total_critical,
        "total_high":       total_high,
        "last_scans": {
            str(mid): {
                "scanned_at":          r.get("scanned_at"),
                "vulnerable_packages": r.get("vulnerable_packages"),
                "risk_level":          r.get("risk_level"),
            }
            for mid, r in _scan_cache.items()
        }
    }


@router.get("/export/excel")
def export_cve_excel(
    request: Request,
    token: str = Query(None),
    db: Session = Depends(get_db)
):
    """Advanced multi-sheet CVE report — full scan history, per-finding
    detail, scan activity log, and remediation guidance. Accepts EITHER a
    normal Authorization header OR a ?token= query param, same dual-auth
    pattern used by report/audit downloads so plain <a href> links work."""
    from auth import decode_token
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    raw_token = token
    if not raw_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.lower().startswith("bearer "):
            raw_token = auth_header[7:]
    if not raw_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(raw_token)
    if not payload.get("sub"):
        raise HTTPException(status_code=401, detail="Invalid token payload")

    scans = db.query(CveScan).order_by(CveScan.scanned_at.desc()).all()
    machines = {m.id: m.hostname for m in db.query(Machine).all()}

    NAVY  = "1E1E2E"
    CYAN  = "00BCD4"
    WHITE = "FFFFFF"
    SEV_COLORS = {
        "CRITICAL": "F44336", "HIGH": "FF9800",
        "MEDIUM": "FFC107", "LOW": "4CAF50", "UNKNOWN": "9E9E9E",
    }

    header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
    header_fill = PatternFill("solid", fgColor=NAVY)
    title_font  = Font(name="Arial", bold=True, color=CYAN, size=16)
    body_font   = Font(name="Arial", size=10)
    thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

    def style_header_row(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def autosize(ws, ncols, min_width=12, max_width=60):
        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            longest = min_width
            for cell in ws[letter]:
                if cell.value:
                    longest = max(longest, min(max_width, len(str(cell.value)) + 2))
            ws.column_dimensions[letter].width = longest

    wb = Workbook()

    # ── Sheet 1: Executive Summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1["A1"] = "JENIX Enterprise — CVE Security Report"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = body_font

    total_scans = len(scans)
    total_findings = sum(s.total_vulns for s in scans)
    total_critical = sum(s.critical for s in scans)
    total_high = sum(s.high for s in scans)
    machines_covered = len(set(s.machine_id for s in scans))
    first_scan = min((s.scanned_at for s in scans), default=None)
    last_scan  = max((s.scanned_at for s in scans), default=None)

    summary_rows = [
        ("Total Scans Run", total_scans),
        ("Machines Covered", machines_covered),
        ("Total Findings (all scans)", total_findings),
        ("Total Critical Findings", total_critical),
        ("Total High Findings", total_high),
        ("First Scan", first_scan.strftime("%Y-%m-%d %H:%M UTC") if first_scan else "N/A"),
        ("Most Recent Scan", last_scan.strftime("%Y-%m-%d %H:%M UTC") if last_scan else "N/A"),
    ]
    r = 4
    ws1.cell(row=r, column=1, value="Metric").font = header_font
    ws1.cell(row=r, column=1).fill = header_fill
    ws1.cell(row=r, column=2, value="Value").font = header_font
    ws1.cell(row=r, column=2).fill = header_fill
    for label, val in summary_rows:
        r += 1
        ws1.cell(row=r, column=1, value=label).font = body_font
        ws1.cell(row=r, column=2, value=val).font = body_font

    r += 2
    ws1.cell(row=r, column=1, value="Per-Machine Latest Risk Snapshot").font = Font(
        name="Arial", bold=True, color=CYAN, size=12)
    r += 1
    headers = ["Machine", "Last Scan", "Risk Level", "Critical", "High", "Total Vulns"]
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=r, column=i, value=h)
    style_header_row(ws1, r, len(headers))
    latest_by_machine = {}
    for s in scans:
        if s.machine_id not in latest_by_machine or s.scanned_at > latest_by_machine[s.machine_id].scanned_at:
            latest_by_machine[s.machine_id] = s
    for mid, s in latest_by_machine.items():
        r += 1
        vals = [machines.get(mid, f"Machine {mid}"), s.scanned_at.strftime("%Y-%m-%d %H:%M"),
                s.risk_level, s.critical, s.high, s.total_vulns]
        for i, v in enumerate(vals, start=1):
            cell = ws1.cell(row=r, column=i, value=v)
            cell.font = body_font
            cell.border = thin_border
    autosize(ws1, 6)

    # ── Sheet 2: Full CVE Detail (all scans) ────────────────────────────
    ws2 = wb.create_sheet("Full CVE Detail")
    headers = ["Scan Date", "Machine", "Package", "Version", "CVE ID", "Severity", "Summary", "Reference URL"]
    for i, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=i, value=h)
    style_header_row(ws2, 1, len(headers))
    row_idx = 1
    for s in scans:
        for f in s.findings:
            row_idx += 1
            vals = [s.scanned_at.strftime("%Y-%m-%d %H:%M"), machines.get(s.machine_id, f"Machine {s.machine_id}"),
                    f.package, f.version, f.cve_id, f.severity, f.summary, f.url]
            for i, v in enumerate(vals, start=1):
                cell = ws2.cell(row=row_idx, column=i, value=v)
                cell.font = body_font
                cell.border = thin_border
            sev_cell = ws2.cell(row=row_idx, column=6)
            sev_cell.fill = PatternFill("solid", fgColor=SEV_COLORS.get(f.severity, "9E9E9E"))
            sev_cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    if row_idx > 1:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx}"
    ws2.freeze_panes = "A2"
    autosize(ws2, len(headers))

    # ── Sheet 3: Scan Activity Log ───────────────────────────────────────
    ws3 = wb.create_sheet("Scan Activity Log")
    headers = ["Scan ID", "Machine", "Triggered By", "Scanned At", "Packages Scanned",
               "Vulnerable Packages", "Total Vulns", "Critical", "High", "Risk Level"]
    for i, h in enumerate(headers, start=1):
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, len(headers))
    for idx, s in enumerate(scans, start=2):
        vals = [s.id, machines.get(s.machine_id, f"Machine {s.machine_id}"), s.triggered_by_name,
                s.scanned_at.strftime("%Y-%m-%d %H:%M:%S"), s.packages_scanned,
                s.vulnerable_packages, s.total_vulns, s.critical, s.high, s.risk_level]
        for i, v in enumerate(vals, start=1):
            cell = ws3.cell(row=idx, column=i, value=v)
            cell.font = body_font
            cell.border = thin_border
    if scans:
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(scans)+1}"
    ws3.freeze_panes = "A2"
    autosize(ws3, len(headers))

    # ── Sheet 4: Remediation Guidance ────────────────────────────────────
    ws4 = wb.create_sheet("Remediation Guidance")
    headers = ["Package", "Current Version", "Highest Severity Found", "CVE Count", "Sample CVE Reference"]
    for i, h in enumerate(headers, start=1):
        ws4.cell(row=1, column=i, value=h)
    style_header_row(ws4, 1, len(headers))

    sev_order = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1, "UNKNOWN": 0}
    pkg_map = {}
    for s in scans:
        for f in s.findings:
            key = (f.package, f.version)
            if key not in pkg_map:
                pkg_map[key] = {"count": 0, "highest": "UNKNOWN", "sample_url": f.url, "sample_cve": f.cve_id}
            pkg_map[key]["count"] += 1
            if sev_order.get(f.severity, 0) > sev_order.get(pkg_map[key]["highest"], 0):
                pkg_map[key]["highest"] = f.severity
                pkg_map[key]["sample_url"] = f.url
                pkg_map[key]["sample_cve"] = f.cve_id

    row_idx = 1
    for (pkg, ver), info in sorted(pkg_map.items(), key=lambda kv: -sev_order.get(kv[1]["highest"], 0)):
        row_idx += 1
        vals = [pkg, ver, info["highest"], info["count"], info["sample_cve"]]
        for i, v in enumerate(vals, start=1):
            cell = ws4.cell(row=row_idx, column=i, value=v)
            cell.font = body_font
            cell.border = thin_border
        sev_cell = ws4.cell(row=row_idx, column=3)
        sev_cell.fill = PatternFill("solid", fgColor=SEV_COLORS.get(info["highest"], "9E9E9E"))
        sev_cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    if row_idx > 1:
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx}"
    ws4.freeze_panes = "A2"
    autosize(ws4, len(headers))

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"jenix_cve_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
